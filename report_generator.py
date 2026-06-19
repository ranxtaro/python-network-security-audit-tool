import json
import shutil
from datetime import datetime
from html import escape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    REPORT_HTML_FILE,
    REPORT_JSON_FILE,
    REPORT_TXT_FILE,
    REPORT_XLSX_FILE,
    TEMPLATE_FILE,
    STYLE_FILE
)


def build_report(network, host_results, risk_summary):
    """Формує загальну структуру звіту."""
    return {
        "час_сканування": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "мережа": str(network),
        "кількість_активних_вузлів": len(host_results),
        "risk_summary": risk_summary,
        "вузли": host_results
    }


def save_reports(report_data):
    """Зберігає звіт у форматах JSON, TXT, HTML та XLSX."""
    save_json_report(report_data)
    save_txt_report(report_data)
    save_html_report(report_data)
    save_excel_report(report_data)


def save_json_report(report_data):
    """Зберігає звіт у форматі JSON."""
    with open(REPORT_JSON_FILE, "w", encoding="utf-8") as json_file:
        json.dump(report_data, json_file, ensure_ascii=False, indent=4)


def save_txt_report(report_data):
    """Зберігає звіт у форматі TXT."""
    summary = report_data["risk_summary"]

    with open(REPORT_TXT_FILE, "w", encoding="utf-8") as txt_file:
        txt_file.write("ЗВІТ З АУДИТУ БЕЗПЕКИ МЕРЕЖІ\n")
        txt_file.write("=" * 40 + "\n")
        txt_file.write(f"Час сканування: {report_data['час_сканування']}\n")
        txt_file.write(f"Мережа: {report_data['мережа']}\n")
        txt_file.write(f"Кількість активних вузлів: {report_data['кількість_активних_вузлів']}\n\n")

        txt_file.write("ПІДСУМОК ЗА РИЗИКАМИ\n")
        txt_file.write("-" * 25 + "\n")
        txt_file.write(f"Високий рівень: {summary['high']}\n")
        txt_file.write(f"Середній рівень: {summary['medium']}\n")
        txt_file.write(f"Низький рівень: {summary['low']}\n")
        txt_file.write(f"Усього ризиків: {summary['total']}\n")
        txt_file.write(f"Вузлів із ризиками: {summary['hosts_with_risks']}\n\n")

        for host in report_data["вузли"]:
            txt_file.write(f"Вузол: {host['ip']}\n")
            txt_file.write(f"Ім'я вузла: {host['hostname']}\n")
            txt_file.write(f"MAC-адреса: {host['mac_address']}\n")
            txt_file.write(f"Виробник: {host['vendor']}\n")
            txt_file.write(f"Тип пристрою: {host['device_type']}\n")

            if host["відкриті_порти"]:
                ports_text = ", ".join(map(str, host["відкриті_порти"]))
                services_text = ", ".join(host["служби"])
                txt_file.write(f"Відкриті порти: {ports_text}\n")
                txt_file.write(f"Служби: {services_text}\n")
                txt_file.write("Банери:\n")

                for item in host["портові_дані"]:
                    txt_file.write(
                        f"  - Порт {item['порт']} | {item['служба']} | {item['банер']}\n"
                    )
            else:
                txt_file.write("Відкритих портів із перевірюваного списку не виявлено\n")

            if host["ризики"]:
                txt_file.write("Ризики:\n")
                for risk in host["ризики"]:
                    port_text = "Н/Д" if risk["порт"] is None else str(risk["порт"])
                    txt_file.write(
                        f"  - {risk['рівень']} | Порт: {port_text} | {risk['опис']}\n"
                    )
            else:
                txt_file.write("Ризики не виявлено\n")

            txt_file.write("\n")


def build_risk_summary_html(summary):
    """Формує HTML-блок із підсумком за ризиками."""
    return f"""
<section class="info-card">
    <h2>Підсумок за ризиками</h2>
    <p><span class="label">Високий рівень:</span> {summary['high']}</p>
    <p><span class="label">Середній рівень:</span> {summary['medium']}</p>
    <p><span class="label">Низький рівень:</span> {summary['low']}</p>
    <p><span class="label">Усього ризиків:</span> {summary['total']}</p>
    <p><span class="label">Вузлів із ризиками:</span> {summary['hosts_with_risks']}</p>
</section>
"""


def build_hosts_html(report_data):
    """Формує HTML-блоки для кожного вузла."""
    blocks = [build_risk_summary_html(report_data["risk_summary"])]

    for host in report_data["вузли"]:
        has_risks = "yes" if host["ризики"] else "no"

        block = [
            f'<section class="host-card" data-risk="{has_risks}">',
            f"<h2>Вузол: {escape(host['ip'])}</h2>",
            f"<p><span class='label'>Ім'я вузла:</span> {escape(host['hostname'])}</p>",
            f"<p><span class='label'>MAC-адреса:</span> {escape(host['mac_address'])}</p>",
            f"<p><span class='label'>Виробник:</span> {escape(host['vendor'])}</p>"
            f"<p><span class='label'>Тип пристрою:</span> {escape(host['device_type'])}</p>"
        ]

        if host["відкриті_порти"]:
            ports_text = ", ".join(map(str, host["відкриті_порти"]))
            services_text = ", ".join(host["служби"])

            block.append(
                f"<p><span class='label'>Відкриті порти:</span> {escape(ports_text)}</p>"
            )
            block.append(
                f"<p><span class='label'>Служби:</span> {escape(services_text)}</p>"
            )

            block.append("<p class='label'>Банери:</p>")
            block.append("<ul>")

            for item in host["портові_дані"]:
                banner_text = f"Порт {item['порт']} | {item['служба']} | {item['банер']}"
                block.append(f"<li>{escape(banner_text)}</li>")

            block.append("</ul>")
        else:
            block.append("<p><span class='label'>Відкриті порти:</span> не виявлено</p>")

        if host["ризики"]:
            block.append('<div class="risk">')
            block.append("<p class='label'>Виявлені ризики:</p>")
            block.append("<ul>")

            for risk in host["ризики"]:
                port_text = "Н/Д" if risk["порт"] is None else str(risk["порт"])
                item_text = f"{risk['рівень']} | Порт: {port_text} | {risk['опис']}"
                block.append(f"<li>{escape(item_text)}</li>")

            block.append("</ul>")
            block.append("</div>")
        else:
            block.append('<div class="no-risk"><p><strong>Ризики не виявлено</strong></p></div>')

        block.append("</section>")
        blocks.append("\n".join(block))

    return "\n".join(blocks)

def copy_style_for_html_report():
    """Копіює CSS-файл у папку static поруч зі звітом HTML."""
    report_dir = Path(REPORT_HTML_FILE).parent
    static_dir = report_dir / "static"
    static_dir.mkdir(parents=True, exist_ok=True)

    target_style_file = static_dir / "style.css"
    shutil.copy2(STYLE_FILE, target_style_file)

def save_html_report(report_data):
    """Зберігає звіт у форматі HTML на основі шаблону."""
    with open(TEMPLATE_FILE, "r", encoding="utf-8") as template_file:
        template = template_file.read()

    html_content = template
    html_content = html_content.replace("__SCAN_TIME__", escape(report_data["час_сканування"]))
    html_content = html_content.replace("__NETWORK__", escape(report_data["мережа"]))
    html_content = html_content.replace(
        "__ACTIVE_HOSTS_COUNT__",
        str(report_data["кількість_активних_вузлів"])
    )
    html_content = html_content.replace("__HOSTS_HTML__", build_hosts_html(report_data))

    copy_style_for_html_report()

    with open(REPORT_HTML_FILE, "w", encoding="utf-8") as html_file:
        html_file.write(html_content)


def apply_header_style(worksheet, row_number):
    """Застосовує стиль до рядка заголовків."""
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def autofit_columns(worksheet):
    """Підбирає ширину стовпців."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def save_excel_report(report_data):
    """Зберігає звіт у форматі Excel."""
    workbook = Workbook()

    # Лист 1 - Підсумок
    summary_sheet = workbook.active
    summary_sheet.title = "Підсумок"

    summary_sheet["A1"] = "ЗВІТ З АУДИТУ БЕЗПЕКИ МЕРЕЖІ"
    summary_sheet["A1"].font = Font(size=14, bold=True)
    summary_sheet.merge_cells("A1:D1")

    summary_sheet["A3"] = "Час сканування"
    summary_sheet["B3"] = report_data["час_сканування"]
    summary_sheet["A4"] = "Мережа"
    summary_sheet["B4"] = report_data["мережа"]
    summary_sheet["A5"] = "Кількість активних вузлів"
    summary_sheet["B5"] = report_data["кількість_активних_вузлів"]

    summary = report_data["risk_summary"]
    summary_sheet["A7"] = "ПІДСУМОК ЗА РИЗИКАМИ"
    summary_sheet["A7"].font = Font(bold=True)

    summary_sheet["A8"] = "Високий рівень"
    summary_sheet["B8"] = summary["high"]
    summary_sheet["A9"] = "Середній рівень"
    summary_sheet["B9"] = summary["medium"]
    summary_sheet["A10"] = "Низький рівень"
    summary_sheet["B10"] = summary["low"]
    summary_sheet["A11"] = "Усього ризиків"
    summary_sheet["B11"] = summary["total"]
    summary_sheet["A12"] = "Вузлів із ризиками"
    summary_sheet["B12"] = summary["hosts_with_risks"]

    # Лист 2 - Вузли
    hosts_sheet = workbook.create_sheet("Вузли")
    hosts_headers = [
        "IP-адреса",
        "Ім'я вузла",
        "MAC-адреса",
        "Виробник",
        "Тип пристрою",
        "Відкриті порти",
        "Служби",
        "Кількість ризиків"
    ]
    hosts_sheet.append(hosts_headers)
    apply_header_style(hosts_sheet, 1)

    for host in report_data["вузли"]:
        hosts_sheet.append([
            host["ip"],
            host["hostname"],
            host["mac_address"],
            host["vendor"],
            host["device_type"],
            ", ".join(map(str, host["відкриті_порти"])) if host["відкриті_порти"] else "Не виявлено",
            ", ".join(host["служби"]) if host["служби"] else "Не виявлено",
            len(host["ризики"])
        ])

    # Лист 3 - Порти_та_ризики
    details_sheet = workbook.create_sheet("Порти_та_ризики")
    details_headers = [
        "IP-адреса",
        "Ім'я вузла",
        "MAC-адреса",
        "Виробник",
        "Тип пристрою",
        "Порт",
        "Служба",
        "Банер",
        "Рівень ризику",
        "Опис ризику"
    ]
    details_sheet.append(details_headers)
    apply_header_style(details_sheet, 1)

    for host in report_data["вузли"]:
        if host["портові_дані"]:
            for port_item in host["портові_дані"]:
                matching_risks = [
                    risk for risk in host["ризики"]
                    if risk["порт"] == port_item["порт"]
                ]

                if matching_risks:
                    for risk in matching_risks:
                        details_sheet.append([
                            host["ip"],
                            host["hostname"],
                            host["mac_address"],
                            host["vendor"],
                            host["device_type"],
                            port_item["порт"],
                            port_item["служба"],
                            port_item["банер"],
                            risk["рівень"],
                            risk["опис"]
                        ])
                else:
                    details_sheet.append([
                        host["ip"],
                        host["hostname"],
                        host["mac_address"],
                        host["vendor"],
                        host["device_type"],
                        port_item["порт"],
                        port_item["служба"],
                        port_item["банер"],
                        "Немає",
                        "Ризик не виявлено"
                    ])
        else:
            details_sheet.append([
                host["ip"],
                host["hostname"],
                host["mac_address"],
                host["vendor"],
                host["device_type"],
                "",
                "",
                "",
                "Немає",
                "Відкритих портів не виявлено"
            ])

    # Стилізація та ширина
    for sheet in workbook.worksheets:
        autofit_columns(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(REPORT_XLSX_FILE)